import openpyxl
import random
from datetime import datetime, timedelta
from copy import copy
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import range_boundaries

# =========================
# FILE PATHS
# =========================
template_path = "data/template.xlsx"
data_path = "data/data.xlsx"
output_path = "output/"

# =========================
# LOAD FILES
# =========================
template_wb = openpyxl.load_workbook(template_path)
data_wb = openpyxl.load_workbook(data_path)

centre_ws = template_wb["Centre Details"]
student_ws = template_wb["Student Registration Details"]
data_ws = data_wb["Sheet1"]

# =========================
# 1️⃣ GET EXACT BLUE COLOR FROM ROWS 5-6
# =========================
# Sample the blue color from row 5, column A
sample_cell = student_ws.cell(row=5, column=1)
blue_color = 'FF0070C0'  # Default fallback

if sample_cell.fill and sample_cell.fill.start_color:
    if hasattr(sample_cell.fill.start_color, 'rgb') and sample_cell.fill.start_color.rgb:
        blue_color = sample_cell.fill.start_color.rgb
    elif hasattr(sample_cell.fill.start_color, 'index') and sample_cell.fill.start_color.index:
        blue_color = sample_cell.fill.start_color.index

print(f"📋 Template blue color: {blue_color}")

# =========================
# 2️⃣ UPDATE CENTRE DETAILS DATES (CUSTOM FORMAT: M/D/YYYY but displays as DD MMM YYYY)
# =========================

today = datetime.today()
start_date = today + timedelta(days=21)
end_date = start_date + timedelta(days=30)

# Update start date - Custom format: shows as "12 Feb 2026" when cell is released
for row in centre_ws.iter_rows():
    for cell in row:
        if cell.value and "Start Date" in str(cell.value):
            date_cell = centre_ws.cell(row=cell.row, column=cell.column + 1)
            if not isinstance(date_cell, openpyxl.cell.cell.MergedCell):
                date_cell.value = start_date.date()
                # This format shows as "12 Feb 2026" when cell is released
                date_cell.number_format = 'DD MMM YYYY'
        if cell.value and "End Date" in str(cell.value):
            date_cell = centre_ws.cell(row=cell.row, column=cell.column + 1)
            if not isinstance(date_cell, openpyxl.cell.cell.MergedCell):
                date_cell.value = end_date.date()
                # This format shows as "12 Feb 2026" when cell is released
                date_cell.number_format = 'DD MMM YYYY'

# =========================
# 3️⃣ READ STUDENT DATA
# =========================
students = []
batch_code = None

for row in data_ws.iter_rows(min_row=2, values_only=True):
    current_batch_code = row[0]
    enrollment = row[2]
    semester = row[3]
    name = row[4]
    
    if batch_code is None and current_batch_code:
        batch_code = str(current_batch_code)
    
    if enrollment and name:
        students.append({
            "batch_code": current_batch_code,
            "enrollment": enrollment,
            "semester": semester,
            "name": name
        })

# =========================
# 4️⃣ CREATE GROUPS (DYNAMIC - MAX 6 PER GROUP)
# =========================
random.shuffle(students)
groups = []
group_size = 6

for i in range(0, len(students), group_size):
    groups.append(students[i:i + group_size])

print(f"\n{'='*50}")
print(f"📊 STUDENT DISTRIBUTION")
print(f"{'='*50}")
print(f"Total Students: {len(students)}")
print(f"Groups created: {len(groups)}")
print(f"Group sizes: {[len(g) for g in groups]}")
print(f"{'='*50}\n")

# =========================
# 5️⃣ PREPARE WORKSHEET (KEEP ROWS 1-6 INTACT)
# =========================
# Unmerge ONLY data area cells (row 7 and below)
merged_ranges_to_remove = []
for merged_range in student_ws.merged_cells.ranges:
    range_str = str(merged_range)
    if '7' in range_str or any(str(i) in range_str for i in range(8, 1000)):
        merged_ranges_to_remove.append(range_str)

for merge_range in merged_ranges_to_remove:
    try:
        student_ws.unmerge_cells(merge_range)
    except:
        pass

# Delete all rows from 7 onwards
if student_ws.max_row >= 7:
    student_ws.delete_rows(7, student_ws.max_row - 6)

# =========================
# 6️⃣ CREATE TEMPLATE ROW AT ROW 7
# =========================
# Create a new row 7 with proper formatting
student_ws.insert_rows(7)

# Set up border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format row 7
for col_idx in range(1, 12):
    cell = student_ws.cell(row=7, column=col_idx)
    cell.border = copy(thin_border)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    if col_idx in [1, 2]:  # A and B columns - center align
        cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = None

# =========================
# 7️⃣ ADD NEW DATA STARTING FROM ROW 7
# =========================
current_row = 7
sr_no = 1

# Group letters mapping
group_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

# Define styles
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal='center', vertical='center')
blue_fill = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
blue_font = Font(bold=True, color='FFFFFF')
blue_alignment = Alignment(horizontal='center', vertical='center')

for group_num, group in enumerate(groups, start=1):
    group_letter = group_letters[group_num - 1] if group_num <= len(group_letters) else str(group_num)
    
    # Store the start row for this group
    group_start_row = current_row
    
    # Add all students in this group
    for student in group:
        # Write values directly to cells
        student_ws.cell(row=current_row, column=1).value = sr_no  # A - Sr. No.
        student_ws.cell(row=current_row, column=2).value = group_letter  # B - Group
        student_ws.cell(row=current_row, column=3).value = student["semester"]  # C
        student_ws.cell(row=current_row, column=4).value = student["batch_code"]  # D
        student_ws.cell(row=current_row, column=5).value = student["enrollment"]  # E
        student_ws.cell(row=current_row, column=6).value = student["name"]  # F
        
        # EMAIL HYPERLINKS - Column G and H
        # Student's eMail-ID
        email_g = "aptechsheh@gmail.com"
        cell_g = student_ws.cell(row=current_row, column=7)
        cell_g.value = email_g
        cell_g.hyperlink = f"mailto:{email_g}"
        cell_g.font = Font(color='0000FF', underline='single')  # Blue, underlined
        
        # Alternate eMail-ID
        email_h = "projects@aptech-metro.com.pk"
        cell_h = student_ws.cell(row=current_row, column=8)
        cell_h.value = email_h
        cell_h.hyperlink = f"mailto:{email_h}"
        cell_h.font = Font(color='0000FF', underline='single')  # Blue, underlined
        
        student_ws.cell(row=current_row, column=9).value = "OV-ACCP Prime-7062-ACE"  # I
        student_ws.cell(row=current_row, column=10).value = "OV-7062-ADSE"  # J
        student_ws.cell(row=current_row, column=11).value = "ACCP Prime-Term 5"  # K
        
        # Apply yellow background to group cell
        cell_b = student_ws.cell(row=current_row, column=2)
        cell_b.fill = copy(yellow_fill)
        cell_b.font = copy(bold_font)
        cell_b.alignment = copy(center_alignment)
        
        sr_no += 1
        current_row += 1
    
    # Store the end row for this group (last student row)
    group_end_row = current_row - 1
    
    # MERGE THE GROUP CELLS (Column B only) from start to end
    if group_start_row < group_end_row:  # Only merge if more than 1 student
        try:
            merge_range = f'B{group_start_row}:B{group_end_row}'
            student_ws.merge_cells(merge_range)
            # Apply yellow formatting to merged cell
            merged_cell = student_ws.cell(row=group_start_row, column=2)
            merged_cell.fill = copy(yellow_fill)
            merged_cell.font = copy(bold_font)
            merged_cell.alignment = copy(center_alignment)
        except:
            pass
    
    # Add TWO blue merged rows AFTER the group ends (EXACTLY like rows 5-6)
    if group_num < len(groups):
        # Add first blue row
        student_ws.insert_rows(current_row)
        for col_idx in range(1, 12):
            cell = student_ws.cell(row=current_row, column=col_idx)
            cell.fill = copy(blue_fill)
            cell.font = copy(blue_font)
            cell.alignment = copy(blue_alignment)
            cell.border = copy(thin_border)
            cell.value = None
        current_row += 1
        
        # Add second blue row
        student_ws.insert_rows(current_row)
        for col_idx in range(1, 12):
            cell = student_ws.cell(row=current_row, column=col_idx)
            cell.fill = copy(blue_fill)
            cell.font = copy(blue_font)
            cell.alignment = copy(blue_alignment)
            cell.border = copy(thin_border)
            cell.value = None
        
        # Merge the two blue rows (A through K) - EXACTLY like rows 5-6
        try:
            student_ws.merge_cells(f'A{current_row-1}:K{current_row}')
        except:
            pass
        
        current_row += 1

# =========================
# 8️⃣ APPLY BORDERS TO ALL DATA ROWS
# =========================
# Apply borders to all data rows
for row in range(7, current_row):
    for col_idx in range(1, 12):
        cell = student_ws.cell(row=row, column=col_idx)
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.border = copy(thin_border)

# =========================
# 9️⃣ GENERATE FILENAME
# =========================
current_date = datetime.now().strftime("%Y%m%d")
invalid_chars = '<>:"/\\|?*'
if batch_code:
    for char in invalid_chars:
        batch_code = batch_code.replace(char, '_')
    batch_code = batch_code[:30]
else:
    batch_code = "UNKNOWN_BATCH"

output_filename = f"eproject_request_{batch_code}_{current_date}.xlsx"
output_full_path = f"{output_path}{output_filename}"

# =========================
# SAVE FILE
# =========================
template_wb.save(output_full_path)

print(f"\n{'='*50}")
print(f"✅ SUCCESS!")
print(f"{'='*50}")
print(f"📁 File: {output_filename}")
print(f"👥 Total Students: {len(students)}")
print(f"👥 Groups: {len(groups)}")
print(f"📊 Group sizes: {[len(g) for g in groups]}")
print(f"🎨 Blue color used: {blue_color}")
print(f"📧 Email hyperlinks: YES (blue, underlined)")
print(f"📅 Date format: DD MMM YYYY (e.g., 12 Feb 2026)")
print(f"🏷️ Batch Code: {batch_code}")
print(f"💾 Saved to: {output_full_path}")
print(f"{'='*50}")